'''
    1- Entrar na planilha e extrair o CPF do cliente
    2- Entrar no site https://consultcpf-devaprender.netlify.app/ e usar o CPF para pesquisar o status de pagamento de uma determinada pessoa
    3- Verificar se o pagamento está "em dia" ou "atrasado"
    4- Se o status do pagamento estiver "em dia", pegar a data e o método de pagamento
    5- Caso o status de pagamento estiver como "atrasado", colocar o status como pendente 
    6- Inserir essas informações (nome, valor, CPF, vencimento, status e caso o pagamento esteja "em dia", adicionar, data pagamento, método pagamento(cartão ou boleto)) em uma nova planilha
    7- Repetir até chegar no ultimo cliente
'''

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

# 1- Entrar na planilha e extrair o CPF do cliente
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

driver = webdriver.Chrome()
driver.get("https://consultcpf-devaprender.netlify.app/")

for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento= linha

    #2- Entrar no site https://consultcpf-devaprender.netlify.app/ e usar o CPF para pesquisar o status de pagamento de uma determinada pessoa

    sleep(5)
    campo_pesquisa = driver.find_element(By.XPATH,"//input[@id='cpfInput']")
    sleep(1)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(1)

    # 3- Verificar se o pagamento está "em dia" ou "atrasado"
    botao_pesquisar = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    botao_pesquisar.click()
    sleep(4)

    #  4- Se o status do pagamento estiver "em dia", pegar a data e o método de pagamento

    status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
    if status.text =='em dia':
        data_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")

        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])

        planilha_fechamento.save('planilha fechamento.xlsx')

    #5- Caso o status de pagamento estiver como "atrasado", colocar o status como pendente 
    else:
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])

        planilha_fechamento.save('planilha fechamento.xlsx')